# streamlit_app.py  (two-step download UX)
import os, math, re, functools
from io import BytesIO
from collections import defaultdict

import streamlit as st
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from mistralai import Mistral
from tenacity import retry, wait_exponential, stop_after_attempt, retry_if_exception_type

# ══════════════════════════════════════
# 🔑  API & model
# ══════════════════════════════════════
API_KEY = st.secrets.get("MISTRAL_API_KEY") or os.getenv("MISTRAL_API_KEY")
MODEL   = "mistral-large-latest"
client  = Mistral(api_key=API_KEY)

# ══════════════════════════════════════
# 🔧  helpers
# ══════════════════════════════════════
def _clean(x):
    if x is None or (isinstance(x, float) and math.isnan(x)): return ""
    return str(x).strip()

def wb_bytes(wb):
    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def wrap(ws):
    for r in ws.iter_rows():
        for c in r:
            c.alignment = Alignment(wrap_text=True,
                                    vertical="top", horizontal="left")

# --- Mistral wrappers ---------------------------------------------------------
@functools.lru_cache(maxsize=1024)
@retry(wait=wait_exponential(multiplier=1,min=2,max=20),
       stop=stop_after_attempt(5),
       retry=retry_if_exception_type(Exception))
def _fmt_uncached(txt):
    res = client.agents.complete(
        agent_id="ag:9d0568a2:20250612:cleaner:12c5f2da",
        messages=[{"role":"user","content":txt}])
    return re.sub(r"[`]+","",res.choices[0].message.content)

def fmt(txt):
    try: return _fmt_uncached(_clean(txt))
    except Exception as e:
        st.error(f"format_text err → {e}"); return _clean(txt)

@functools.lru_cache(maxsize=1024)
@retry(wait=wait_exponential(multiplier=1,min=2,max=20),
       stop=stop_after_attempt(5),
       retry=retry_if_exception_type(Exception))
def _manu_uncached(txt):
    res = client.chat.complete(
        model=MODEL,
        messages=[{"role":"user",
                   "content":("Extract manufacturer names separated by hyphen - "
                              "plain list only\ncontent: "+txt)}])
    return res.choices[0].message.content

def manu(txt):
    try: return _manu_uncached(_clean(txt))
    except Exception as e:
        st.error(f"manu_name err → {e}"); return ""

# --- PDF helpers --------------------------------------------------------------
def pdf_txt(pdf):
    return "".join(p.extract_text() for p in PdfReader(pdf).pages)

def pdf_body(pdf):
    return re.sub(r"(REQUEST FOR QUOTATION[\s\S]*?RFQ Number \d+)",
                  "", pdf_txt(pdf))

def parse_pdf(body, rfq):
    rfx = re.search(r"RFQ Number (\d+)", rfq)
    rfx_no = rfx.group(1) if rfx else "Unknown"

    pat = re.compile(r"(\d{5}) (\w?12\d{10}) (\d+(?:\.\d+)?)\s*(\w+) .*?"
                     r"(\d{2}\.\d{2}\.\d{4})", re.DOTALL)
    short = re.findall(r"Short Text :(.*?)\n", body, re.DOTALL)
    po    = re.findall(r"PO Material Text :(.*?)Agreement / LineNo.", body,
                       re.DOTALL)

    data=[]
    for i,m in enumerate(pat.findall(body)):
        mat = m[1] if m[1].startswith(("B12","12","B16","15")) else ""
        data.append({"RFx Number":rfx_no,"RFx Item No":m[0],"PR Item No":"",
                     "Material No":mat,
                     "Description":short[i] if i<len(short) else "",
                     "PO Text":po[i] if i<len(po) else "",
                     "QTY":m[2],"UOM":m[3]})
    return data

# ══════════════════════════════════════
# 🖥️  Streamlit UI
# ══════════════════════════════════════
st.set_page_config(page_title="Data Processor", layout="wide",
                   initial_sidebar_state="collapsed")
col1,col2,col3=st.columns([2,2,1])

# ───────────────────────────────────────
# 1️⃣  Excel workflow
# ───────────────────────────────────────
with col1:
    st.subheader("🗃️ Excel Data → Upload / Final")
    techno = st.file_uploader("Techno-Commercial (.xls)",type=["xls"],
                              key="xls")
    upl_tpl = st.file_uploader("Upload template (.xlsx)",type=["xlsx"],
                               key="upl_tpl") or "upload file - HTS.xlsx"
    fin_tpl = st.file_uploader("Final Sheet template (.xlsx)",type=["xlsx"],
                               key="fin_tpl") or "FINAL SHEET.xlsx"
    suffix  = st.text_input("Name suffix", key="suf")

    # ---------- STEP 1 : Upload ----------
    if st.button("🚀 Generate Upload file", key="btn_upl") and techno and suffix:
        try:
            rfx_no = re.search(r"\d+", techno.name).group()
            xls    = pd.ExcelFile(techno)
            need   = {'Description','InternalNote','Quantity','Unit of Measure'}
            sheet  = next((s for s in xls.sheet_names
                          if need.issubset(set(pd.read_excel(xls,s,nrows=1).columns))),
                          None)
            if not sheet: st.error("Required cols missing"); st.stop()
            df = pd.read_excel(techno, sheet_name=sheet, keep_default_na=False)
            valid = df[(df['Description'].str.strip().str.lower()!='item or lot description') &
                       df['Quantity'].astype(str).str.strip().ne('') &
                       df['Unit of Measure'].astype(str).str.strip().ne('') &
                       (df['Unit of Measure'].str.strip().str.lower()!='unit of measure')]

            wb_u = load_workbook(upl_tpl); ws=wb_u.active
            for r in ws.iter_rows(min_row=2,max_row=ws.max_row):
                for c in r:c.value=None
            row,item=2,10
            for _,rec in valid.iterrows():
                ws[f"A{row}"]=rfx_no
                ws[f"B{row}"]=item
                ws[f"E{row}"]=rec['Description']
                ws[f"H{row}"]=rec['Unit of Measure']
                ws[f"G{row}"]=rec['Quantity']
                ws[f"F{row}"]=rec['InternalNote']
                ws[f"I{row}"]=rec.get('Number','')
                item+=10; row+=1
            wrap(ws)

            st.session_state["excel_valid"] = valid  # cache DF for step 2
            st.session_state["excel_upload_bytes"] = wb_bytes(wb_u)
            st.session_state["excel_suffix"] = suffix
            st.success("Upload ready ✔️")
            st.experimental_rerun()
        except Exception as e:
            st.error(f"❌ {e}")

    # download Upload button
    if "excel_upload_bytes" in st.session_state:
        st.download_button("📥 Download Upload file",
                           st.session_state["excel_upload_bytes"],
                           file_name=f"upload file - {st.session_state['excel_suffix']}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_upl_excel")

        # ---------- STEP 2 : Final ----------
        if st.button("🚀 Generate FINAL SHEET", key="btn_fin"):
            try:
                valid = st.session_state["excel_valid"]
                wb_f  = load_workbook(fin_tpl); ws=wb_f.active
                for r in ws.iter_rows(min_row=2,max_row=ws.max_row):
                    for c in r:c.value=None
                row,item=2,10
                for _,rec in valid.iterrows():
                    ws[f"A{row}"]=item
                    ws[f"B{row}"]=rec['Description']
                    ws[f"C{row}"]=rec['Quantity']
                    ws[f"D{row}"]=rec['Unit of Measure']
                    po=rec['InternalNote']
                    ws[f"E{row}"]=fmt(po)
                    ws[f"G{row}"]=manu(po)
                    item+=10; row+=1
                wrap(ws)
                st.session_state["excel_final_bytes"] = wb_bytes(wb_f)
                st.success("FINAL SHEET ready ✔️")
                st.experimental_rerun()
            except Exception as e:
                st.error(f"❌ {e}")

    if "excel_final_bytes" in st.session_state:
        st.download_button("📥 Download FINAL SHEET",
                           st.session_state["excel_final_bytes"],
                           file_name=f"FINAL SHEET - {st.session_state['excel_suffix']}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_fin_excel")

# ───────────────────────────────────────
# 2️⃣  PDF workflow
# ───────────────────────────────────────
with col2:
    st.subheader("📑 PDF → Upload / Final")
    pdf = st.file_uploader("RFQ PDF",type=["pdf"], key="pdf")
    raw_tpl = st.file_uploader("Raw template (.xlsx)",type=["xlsx"],
                               key="raw_tpl") or "raw_template.xlsx"
    hts_tpl = st.file_uploader("HTS template (.xlsx)",type=["xlsx"],
                               key="hts_tpl") or "upload file - HTS.xlsx"
    fin_tpl_p = st.file_uploader("Final Sheet template (.xlsx)",type=["xlsx"],
                                 key="fin_tpl_p") or "FINAL SHEET.xlsx"
    hts_no = st.text_input("HTS number", key="hts_no")

    # ---------- STEP 1 : Upload ----------
    if st.button("🚀 Generate Upload from PDF", key="btn_pdf_upl") and pdf and hts_no:
        try:
            data = parse_pdf(pdf_body(pdf), pdf_txt(pdf))
            wb_up = load_workbook(hts_tpl); ws=wb_up.active
            for r in ws.iter_rows(min_row=2,max_row=ws.max_row):
                for c in r:c.value=None
            row=2
            for rec in data:
                for col,let in zip(
                    ["RFx Number","RFx Item No","PR Item No","Material No",
                     "Description","PO Text","QTY","UOM"],
                    list("ABCD")+list("EFGH")):
                    ws[f"{let}{row}"]=rec[col]
                row+=1
            wrap(ws)
            st.session_state["pdf_data"]=data
            st.session_state["pdf_upload_bytes"]=wb_bytes(wb_up)
            st.session_state["pdf_hts"]=hts_no
            st.success("Upload ready ✔️")
            st.experimental_rerun()
        except Exception as e:
            st.error(f"❌ {e}")

    if "pdf_upload_bytes" in st.session_state:
        h=st.session_state["pdf_hts"]
        st.download_button("📥 Download Upload file",
                           st.session_state["pdf_upload_bytes"],
                           file_name=f"upload file - {h}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_pdf_upl")

        # ---------- STEP 2 : Final ----------
        if st.button("🚀 Generate FINAL SHEET", key="btn_pdf_fin"):
            try:
                data=st.session_state["pdf_data"]
                wb_fin=load_workbook(fin_tpl_p); wsf=wb_fin.active
                for r in wsf.iter_rows(min_row=2,max_row=wsf.max_row):
                    for c in r:c.value=None
                row=2
                for rec in data:
                    wsf[f"A{row}"]=rec['RFx Item No']
                    wsf[f"B{row}"]=rec['Description']
                    wsf[f"C{row}"]=rec['QTY']
                    wsf[f"D{row}"]=rec['UOM']
                    wsf[f"E{row}"]=fmt(rec['PO Text'])
                    wsf[f"G{row}"]=manu(rec['PO Text'])
                    row+=1
                wrap(wsf)
                st.session_state["pdf_final_bytes"]=wb_bytes(wb_fin)
                st.success("FINAL SHEET ready ✔️")
                st.experimental_rerun()
            except Exception as e:
                st.error(f"❌ {e}")

    if "pdf_final_bytes" in st.session_state:
        h=st.session_state["pdf_hts"]
        st.download_button("📥 Download FINAL SHEET",
                           st.session_state["pdf_final_bytes"],
                           file_name=f"FINAL SHEET - {h}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_pdf_fin")

# ───────────────────────────────────────
# 3️⃣  HTS Cleaner & List Maker
# ───────────────────────────────────────
with col3:
    st.subheader("🧹 HTS Cleaner")
    hts_up = st.file_uploader("upload file – HTS.xlsx",type=["xlsx"], key="hts_up")
    fin_tpl_opt = st.file_uploader("Final Sheet template (opt)",
                                   type=["xlsx"], key="fin_tpl_opt")\
                  or "FINAL SHEET.xlsx"
    if st.button("🚀 Clean HTS", key="btn_clean") and hts_up:
        try:
            wb_up=load_workbook(hts_up)
            wb_fin=load_workbook(fin_tpl_opt); wsf=wb_fin.active
            for r in wsf.iter_rows(min_row=2,max_row=wsf.max_row):
                for c in r:c.value=None
            up_ws=wb_up.active
            row=2
            for r in up_ws.iter_rows(min_row=2,max_row=up_ws.max_row):
                if not any(c.value for c in r): continue
                wsf[f"A{row}"]=r[1].value
                wsf[f"B{row}"]=r[4].value
                wsf[f"C{row}"]=r[6].value
                wsf[f"D{row}"]=r[7].value
                po=r[5].value or ""
                wsf[f"E{row}"]=fmt(po)
                wsf[f"G{row}"]=manu(po)
                row+=1
            wrap(wsf)
            st.download_button("📥 Download cleaned FINAL SHEET",
                               wb_bytes(wb_fin),
                               file_name="FINAL SHEET - cleaned.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_hts_clean")
        except Exception as e:
            st.error(f"❌ {e}")

    # List maker unchanged …
